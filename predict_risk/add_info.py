import openpyxl
def insert_into_dataset(predictions):
    wb = openpyxl.load_workbook('D:/Projects/Heart-disease-prediction-master/predict_risk/classified_dataset.xlsx')
    sheet1=wb.active
    max_row = sheet1.max_row  # Get max row of first sheet
    max_col = sheet1.max_column  # Get max column of first sheet
    row=max_row+1
    for col_num in range(1, max_col + 1):
        cell1 = sheet1.cell(row=row, column=col_num)
        cell1.value = predictions[col_num-1]
    wb.save('D:/Projects/Heart-disease-prediction-master/predict_risk/classified_dataset.xlsx')